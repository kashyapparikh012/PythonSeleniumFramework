import openpyxl


class HomePageData:

    test_HomePage_Data = [{"firstname": "Kashyap", "email": "kyp1395@gmail.com", "password": "Test@123", "gender": "Male"}, {"firstname": "Riya", "email": "rkp1395@gmail.com", "password": "Retest@456", "gender": "Female"}]

    @staticmethod
    def getHomepageData(test_case_name):
        Dict = {}  # Dict: <class 'dict'>: {'firstname': 'Meet', 'email': 'meet@gmail.com'}
        book = openpyxl.load_workbook("C:/Users/kashy/PycharmProjects/PythonSelFramework/PythonDemo.xlsx")
        sheet = book.active

        for i in range(1, sheet.max_row + 1):  # To get rows
            if sheet.cell(row=i, column=1).value == test_case_name:
                for j in range(2, sheet.max_column + 1):  # To get columns
                    # Dict["firstname"] = "Meet"
                    Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
        return Dict
