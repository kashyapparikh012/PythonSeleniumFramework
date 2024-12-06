import openpyxl
Dict = {}   # Dict: <class 'dict'>: {'firstname': 'Meet', 'email': 'meet@gmail.com'}
book = openpyxl.load_workbook("C:/Users/kashy/PycharmProjects/PythonSelFramework/PythonDemo.xlsx")
sheet = book.active
# cell = sheet.cell(row=1, column=2)
# print(cell.value)
#
# sheet.cell(row=2, column=2).value = "Meet"
# print(sheet.cell(row=2, column=2).value)
#
# print(sheet.max_row)
# print(sheet.max_column)
#
# print(sheet['A3'].value)

for i in range(1, sheet.max_row+1): # To get rows
    if sheet.cell(row=i, column=1).value == "TestCase1":
        for j in range(2, sheet.max_column+1):  # To get columns
            #Dict["firstname"] = "Meet"
            Dict[sheet.cell(row=1, column=j).value] = sheet.cell(row=i, column=j).value
print(Dict)
